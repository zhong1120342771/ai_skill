"""
把两个本地xlsx的数据写入飞书表格
  莫斯科保卫战周报260608.xlsx   → sheet名加 _大盘 后缀
  莫斯科保卫战周报补充数据格式.xlsx → sheet名加 _分业务 后缀
"""

import json, subprocess, time, math
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

STOKEN = "HsL6sUaUuhcbhWtKSplcSym4npg"
CHUNK  = 4000   # 每次写入行数上限（飞书单次write上限约5000行）

XLSX_TASKS = [
    ('/Users/zhongmengting/Downloads/莫斯科保卫战周报260608.xlsx',          '_大盘'),
    ('/Users/zhongmengting/Downloads/莫斯科保卫战周报补充数据格式.xlsx',    '_分业务'),
]

# 跳过图片专用sheet（无有效数据，另外处理）
SKIP_SHEETS = {'图片', '分业务图表'}

# ── 工具函数 ──────────────────────────────────────────────────────────────────

def run(cmd, silent=False):
    r = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    if not silent:
        if r.returncode != 0:
            print(f'  [ERR] {r.stderr[:300]}')
        return r.stdout.strip()
    return r.stdout.strip()

def lark_create_sheet(title, index):
    """新建sheet，返回 sheet_id"""
    cmd = (f'lark-cli sheets +create-sheet '
           f'--spreadsheet-token {STOKEN} '
           f'--title "{title}" '
           f'--index {index} '
           f'--as user')
    out = run(cmd)
    try:
        d = json.loads(out)
        sid = d['data']['sheet']['sheet_id']
        print(f'  Created sheet: {title} -> {sid}')
        return sid
    except Exception as e:
        print(f'  [ERR] create_sheet failed: {e}\n  raw: {out[:200]}')
        return None

def lark_rename_sheet(sheet_id, new_title):
    cmd = (f'lark-cli sheets +update-sheet '
           f'--spreadsheet-token {STOKEN} '
           f'--sheet-id {sheet_id} '
           f'--title "{new_title}" '
           f'--as user')
    run(cmd, silent=True)

def cell_value(v):
    """把 openpyxl 的值转为 JSON 安全的 Python 类型"""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        if math.isnan(v) or math.isinf(v):
            return ""
        return v
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    return str(v)

def col_letter(n):
    """1-indexed col number -> Excel column letter (A, B, ..., Z, AA, ...)"""
    s = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def write_rows_to_sheet(sheet_id, rows):
    """
    rows: list of list
    必须在 range 中指定完整的 A1:ZZ{N} 范围。
    使用临时文件传递 values 避免 shell 引号冲突。
    """
    if not rows:
        return
    import tempfile, os
    total    = len(rows)
    n_cols   = max(len(r) for r in rows)
    col_end  = col_letter(n_cols)
    written  = 0
    row_start = 1
    while written < total:
        chunk = rows[written: written + CHUNK]
        padded = [r + [''] * (n_cols - len(r)) for r in chunk]
        row_end   = row_start + len(chunk) - 1
        range_str = f"{sheet_id}!A{row_start}:{col_end}{row_end}"
        values_json = json.dumps(padded, ensure_ascii=False)
        # 写临时文件，避免 shell 引号问题
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json',
                                        delete=False, encoding='utf-8') as f:
            f.write(values_json)
            tmp_path = f.name
        try:
            values_arg = open(tmp_path).read()
            cmd = ['lark-cli', 'sheets', '+write',
                   '--spreadsheet-token', STOKEN,
                   '--range', range_str,
                   '--values', values_arg,
                   '--as', 'user']
            r = subprocess.run(cmd, capture_output=True, text=True)
            out = r.stdout.strip()
            try:
                d = json.loads(out)
                if not d.get('ok'):
                    print(f'    [WARN] write chunk failed: {out[:300]}')
                else:
                    print(f'    Chunk written: rows {row_start}-{row_end}, '
                          f'updatedCells={d["data"].get("updatedCells")}')
            except:
                print(f'    [WARN] parse response: {out[:100]} | err: {r.stderr[:100]}')
        finally:
            os.unlink(tmp_path)
        written   += len(chunk)
        row_start += len(chunk)
        if written < total:
            time.sleep(0.4)

def process_xlsx(path, suffix, existing_sheet_map, sheet_index_start):
    """
    existing_sheet_map: {title -> sheet_id} 已存在的sheet
    """
    wb = load_workbook(path, data_only=True)
    idx = sheet_index_start
    for sh_name in wb.sheetnames:
        if sh_name in SKIP_SHEETS:
            print(f'  Skipping {sh_name} (image-only)')
            continue

        ws = wb[sh_name]
        target_title = sh_name + suffix
        print(f'\n  Processing: {sh_name} -> [{target_title}]')

        rows = []
        for row in ws.iter_rows(values_only=True):
            converted = [cell_value(c) for c in row]
            rows.append(converted)
        # 去掉末尾全空行
        while rows and not any(v != "" for v in rows[-1]):
            rows.pop()

        if not rows:
            print(f'    No data, skipping.')
            continue

        print(f'    Rows: {len(rows)}, Cols: {len(rows[0]) if rows else 0}')

        # 复用已有sheet 或 新建
        if target_title in existing_sheet_map:
            sid = existing_sheet_map[target_title]
            print(f'  Reusing sheet: {target_title} -> {sid}')
        else:
            sid = lark_create_sheet(target_title, idx)
            if sid is None:
                continue
            idx += 1
            time.sleep(0.3)

        write_rows_to_sheet(sid, rows)
        print(f'    Done: {len(rows)} rows.')
        time.sleep(0.5)

    return idx

# ── 主流程 ────────────────────────────────────────────────────────────────────

def main():
    out = run(f'lark-cli sheets +info --spreadsheet-token {STOKEN} --as user', silent=True)
    d = json.loads(out)
    existing = d['data']['sheets']['sheets']
    print(f'Existing sheets: {[s["title"] for s in existing]}')
    default_id  = existing[0]['sheet_id']   # Sheet1
    # 已有的 title->id 映射（用于复用，不重复创建）
    sheet_map = {s['title']: s['sheet_id'] for s in existing}

    idx = len(existing)  # 新sheet从已有数量之后开始

    for path, suffix in XLSX_TASKS:
        print(f'\n=== {path} (suffix: {suffix}) ===')
        idx = process_xlsx(path, suffix, sheet_map, idx)

    # 删除默认 Sheet1（如果还存在）
    if 'Sheet1' in sheet_map:
        print(f'\nDeleting default Sheet1 ({default_id})...')
        cmd = ['lark-cli', 'sheets', '+delete-sheet',
               '--spreadsheet-token', STOKEN,
               '--sheet-id', default_id,
               '--yes',
               '--as', 'user']
        r = subprocess.run(cmd, capture_output=True, text=True)
        print(r.stdout[:200] or r.stderr[:200])

    print('\n✅ All done.')
    print(f'URL: https://zhuanspirit.feishu.cn/sheets/{STOKEN}')

if __name__ == '__main__':
    main()
